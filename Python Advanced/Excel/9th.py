import openpyxl as xl

wb = xl.load_workbook("FSI.xlsx")
ws = wb["Sheet1"]

print(ws.max_column)
ws.append(["Lokesh",  2023,  "1st",  111.89999999999999,  10,  9,  8.7,  8.6,  9.1,  9.5,  9.6,  9.8 , 9,  9.5,  10,  9.1 ])

wb.save("FSI.xlsx")

# print(ws[ws.max_row-1])
last_cell = ws[ws.max_row]

for cell in last_cell:
    print(cell.value, end=" ")