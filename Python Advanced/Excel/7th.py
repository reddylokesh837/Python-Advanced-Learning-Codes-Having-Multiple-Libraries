import openpyxl as xl

wb = xl.load_workbook("FSI.xlsx")

ws = wb['Sheet1']

emp_id_to_match = 'Somalia'

match = 0
cells= None

for row in range(1, ws.max_row+1):
    if ws.cell(row,1).value == emp_id_to_match:
        cells = ws[row]
        match =1
        break

if match:
    for cell in cells:
        print(cell.value, end="  ")
else:
    print("Employee detail does not exist")