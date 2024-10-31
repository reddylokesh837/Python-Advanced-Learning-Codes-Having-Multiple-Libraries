import openpyxl as xl

wb = xl.load_workbook("FSI.xlsx")
ws = wb['Sheet1']

cell1 = ws.cell(1,1)
cell2 = ws['A1']
print(cell1.value)
print(cell2.value)
