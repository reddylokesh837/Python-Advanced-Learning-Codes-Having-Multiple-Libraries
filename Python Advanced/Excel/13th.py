import openpyxl as xl

wb = xl.Workbook()


wb.create_sheet(title="Students", index=0)

print(wb.sheetnames)
wb.remove_sheet(wb['Sheet'])
print(wb.sheetnames)

ws = wb['Students']
ws['A1'] = "Student_ID"
ws['B1']= "Student_Name"
ws['C1'] = "Marks1"
ws['D1'] = "Marks2"
ws['E1'] = "Marks3"

ws.append([101, "David", 45, 56, 67])
ws.append([102, "John",67,76,63])
ws.append([103, "Mark",84,82,93])
ws.append([104,"Andrew", 94,59,69])


print(ws.max_row, ws.max_column)


def read_data():
    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            print(ws.cell(row,col).value, end=" ")

        print()

read_data()



marks_col1= 3
marks_col2 = 4
marks_col3 = 5
total_marks_col = 6
ws.cell(row=1, column=total_marks_col).value = "Total Marks"

def total_marks():
    for row in range(2, ws.max_row+1):
        mark1 = ws.cell(row, marks_col1).value or 0
        mark2 = ws.cell(row, marks_col2).value or 0
        mark3 = ws.cell(row, marks_col3).value or 0
        total_mark = mark1+mark2+mark3

        ws.cell(row, total_marks_col).value = total_mark


total_marks()
wb.save("Students_updated.xlsx")

print("Data updated successfully")
















