import openpyxl as xl

wb = xl.load_workbook("FSI.xlsx")

ws = wb['Sheet1']

emp_id_to_match = "Somalia"

cells=None


for col in range(1, ws.max_column+1):
    if ws.cell(1, col).value == "Country":
        for row in range(2, ws.max_row+1):
            if ws.cell(row, col).value == emp_id_to_match:
                cells = ws[row]
        break

for cell in cells:
    print(cell.value, end="  ")