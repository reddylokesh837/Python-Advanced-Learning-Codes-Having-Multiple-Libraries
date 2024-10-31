import openpyxl as xl

wb = xl.load_workbook("Students_updated.xlsx")
ws = wb['Students']
print(ws.max_column)



def update_data():
    student_id = int(input("Enter student id: "))
    marks = int(input("Enter marks: "))

    for row in range(2, ws.max_row+1):
        if ws.cell(row, 1).value== student_id:
            ws.cell(row, 3).value = marks

    wb.save("Students_updated.xlsx")


update_data()