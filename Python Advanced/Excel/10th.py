import openpyxl as xl

wb = xl.load_workbook("FSI.xlsx")
ws = wb["Sheet1"]




name = "Lokesh"

for row in range(1, ws.max_row+1):
    if ws.cell(row,1).value == name:
        ws.cell(row,1).value = "Reddy"


wb.save("FSI.xlsx")

changed_column = ws[ws.max_row]

print(changed_column[0].value)
