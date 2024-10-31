import openpyxl as xl

# wb = xl.Workbook()

# print(wb.sheetnames)

# wb.create_sheet(title="Sheet1", index=0)
# print(wb.sheetnames)
# wb.remove(wb['Sheet'])
# print(wb.sheetnames)

# ws= wb['Sheet1']

# ws['A1']="Name"
# ws['B1'] = "Salary"
# ws.append(["Lokesh", 20000])
# ws.append(['Reddy' , 30000])

# wb.save("Emp.xlsx")


wb = xl.load_workbook("Emp.xlsx")
ws=wb['Sheet1']
print(ws.max_row)
for row in range(1, ws.max_row+1):
    for col in range(1, ws.max_column+1):
        print(ws.cell(row, col).value, end=" ")
