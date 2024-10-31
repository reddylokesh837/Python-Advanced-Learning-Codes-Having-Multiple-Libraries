import openpyxl as xl

wb = xl.load_workbook("FSI.xlsx")

ws = wb['Sheet1']


print("Cell objects of row 2")
print(ws[2])

print("\nCell objects of column B: ")
print(ws['B'])

print("\n")

print("Accessing cell values row-wise: \n")
for row in ws[2]:
    print(row.value, end="  ")

print("\nValues of all the columns of B: ")
for col in ws["B"]:
    print(col.value)