import openpyxl as xl


wb = xl.load_workbook("Students_updated.xlsx")

ws = wb['Students']

print(wb.sheetnames)
print(ws.max_row)
print(ws.max_column)


def read_data():
    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            print(ws.cell(row, col).value, end=" ")
        print()





def remove_cols():
    new_wb = xl.Workbook()
    new_ws = new_wb.create_sheet(title="Students", index=0)


    mark1_col = 3
    mark2_col=4
    mark3_col=5


    for row in range(1, ws.max_row+1):
        mark1_value = ws.cell(row, mark1_col).value or 0
        mark2_value = ws.cell(row, mark2_col).value or 0
        mark3_value = ws.cell(row, mark3_col).value or 0

        new_ws.cell(row, column=1).value = mark1_value
        new_ws.cell(row, column=2).value = mark2_value
        new_ws.cell(row, column=3).value = mark3_value


    new_wb.save("Students_updated_new.xlsx")


read_data()
print(ws.max_column)

remove_cols()
ws.delete_cols(3,3)

wb.save("Students_updated.xlsx")
print(ws.max_column)