import openpyxl as xl


wb = xl.load_workbook("FSI.xlsx")

print(type(wb))
print(wb.active)
print(wb.sheetnames)
print(wb.worksheets)
ws = wb['Sheet1']

print(type(ws))


cell1= ws.cell(1,2)
cell2 = ws['A2']

print(cell1.value)
print(cell2.value)
